import io
from datetime import datetime
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_user, logout_user, current_user, login_required
from urllib.parse import urlparse
from app import app, db
from models import Project, Employee, ProjectStaff, HoursEntry, User, Company
from forms import ProjectForm, EmployeeForm, ProjectStaffForm, HoursEntryForm, SignupForm, LoginForm, CommissionReportForm, DateRangeForm
from utils import get_paginated_query, is_admin_email
from sqlalchemy.orm import joinedload, subqueryload, contains_eager
from sqlalchemy.exc import IntegrityError
from sqlalchemy import or_, func, and_
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _make_excel_response(wb, filename):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _style_header_row(ws):
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2c32f2', end_color='2c32f2', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
# Authentication routes
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if current_user.is_authenticated:
        app.logger.info(f"Signup attempt while already logged in as {current_user.email}")
        return redirect(url_for('index'))
    
    form = SignupForm()
    if form.validate_on_submit():
        app.logger.debug("Signup form submitted")
        # Check if username or email already exists
        existing_user = User.query.filter((User.email == form.email.data)).first()
        if existing_user:
            app.logger.warning(f"Signup failed: Email {form.email.data} already exists")
            flash('Username or email already exists', 'error')
            return render_template('auth/signup.html', form=form)
        if is_admin_email(form.email.data):
            app.logger.info(f"Admin signup detected for email: {form.email.data}")
            company_id = 5
            company = Company.query.filter_by(id=company_id).first()
        else: # Check if company already exists
            company = Company.query.filter_by(name=form.company_name.data).first()
            if not company:
                company = Company(name=form.company_name.data)
                db.session.add(company)
                db.session.flush()
                app.logger.info(f"Created new company: {company.name}")
        # Create user
        user = User(
            username=form.name.data,
            email=form.email.data,
            company_id=company.id,
            is_owner=True  # First user of company becomes owner
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        app.logger.info(f"New user registered: {user.email} (Company: {company.name})")
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))
    
    app.logger.debug("Rendering signup form")
    return render_template('auth/signup.html', form=form)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        app.logger.info(f"Login attempt while already logged in as {current_user.email}")
        return redirect(url_for('index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        app.logger.debug("Login form submitted")
        user = User.query.filter_by(email=form.username.data).first()

        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data)
            app.logger.info(f"User logged in: {user.email}")
            next_page = request.args.get('next')
            if not next_page or urlparse(next_page).netloc != '':
                next_page = url_for('index')
            return redirect(next_page)

        app.logger.warning(f"Failed login attempt for email: {form.username.data}")
        flash('Invalid username or password', 'error')
    
    app.logger.debug("Rendering login form")
    return render_template('auth/login.html', form=form)


@app.route('/logout')
@login_required
def logout():
    app.logger.info(f"User logged out: {current_user.email}")
    logout_user()
    return redirect(url_for('login'))
@app.route('/', methods=['GET', 'POST'])
@login_required
def index():
    app.logger.debug(f"Index page accessed by {current_user.email}")

    form = CommissionReportForm()
    form.set_employee_choices(current_user.company_id)

    report_data = []
    employee = None
    direct_commission = 0.0
    override_commission = 0.0

    if form.validate_on_submit():
        employee_id = form.employee_id.data
        date_from = form.date_from.data
        date_to = form.date_to.data

        app.logger.info(
            f"Commission report requested by {current_user.email} "
            f"for employee_id={employee_id}, date range={date_from} to {date_to}"
        )

        employee = Employee.query.filter_by(
            id=employee_id, company_id=current_user.company_id
        ).first_or_404()

        # Only fetch projects this employee is actually assigned to, scoped to company
        projects = Project.query.options(
            joinedload(Project.project_staff).joinedload(ProjectStaff.employee)
        ).join(ProjectStaff).filter(
            Project.company_id == current_user.company_id,
            ProjectStaff.employee_id == employee_id
        ).all()

        project_ids = [p.id for p in projects]

        # Single query for all hours across all relevant projects in date range
        all_project_hours = HoursEntry.query.filter(
            HoursEntry.project_id.in_(project_ids),
            HoursEntry.date >= date_from,
            HoursEntry.date <= date_to
        ).all() if project_ids else []

        app.logger.debug(
            f"Fetched {len(all_project_hours)} hours entries across "
            f"{len(project_ids)} projects for employee_id={employee_id}"
        )

        # Map: project_id -> employee_id -> total hours worked
        project_emp_hours = {}
        for entry in all_project_hours:
            project_emp_hours.setdefault(entry.project_id, {})
            project_emp_hours[entry.project_id][entry.employee_id] = (
                project_emp_hours[entry.project_id].get(entry.employee_id, 0) + entry.hours_worked
            )

        for project in projects:
            staff_map = {ps.employee_id: ps for ps in project.project_staff}
            staff = staff_map.get(employee_id)
            if not staff:
                continue

            emp = employee
            emp_hours_map = project_emp_hours.get(project.id, {})
            hours = emp_hours_map.get(employee_id, 0)
            revenue = hours * (staff.hourly_rate or 0)

            # Total project revenue: sum across all staff using their rates
            total_project_revenue = sum(
                emp_hours_map.get(eid, 0) * (ps.hourly_rate or 0)
                for eid, ps in staff_map.items()
            )

            # Associate-specific aggregates for non-associate commission basis
            total_assoc_hours = sum(
                emp_hours_map.get(eid, 0)
                for eid, ps in staff_map.items()
                if ps.employee and ps.employee.role.lower() == 'associate'
            )
            total_assoc_revenue = sum(
                emp_hours_map.get(eid, 0) * (ps.hourly_rate or 0)
                for eid, ps in staff_map.items()
                if ps.employee and ps.employee.role.lower() == 'associate'
            )

            # Commission calculation with correct display hours/revenue per role
            role = emp.role.lower()
            if role == 'associate':
                direct_comm = revenue * (staff.commission_percentage or 0) / 100
                override_comm = 0
                display_hours = hours
                display_revenue = revenue
                note = "Associate: own hours commission"
            elif role == 'director':
                direct_comm = revenue * (staff.commission_percentage or 0) / 100
                override_comm = (total_project_revenue * employee.override_percentage / 100) if employee.override_percentage else 0
                display_hours = sum(emp_hours_map.get(eid, 0) for eid in staff_map)
                display_revenue = total_project_revenue
                note = "Director: direct + 2% override"
            else:
                direct_comm = total_assoc_revenue * (staff.commission_percentage or 0) / 100
                override_comm = 0
                display_hours = total_assoc_hours
                display_revenue = total_assoc_revenue
                note = "Other: % of associate revenue"

            app.logger.debug(
                f"Project {project.name} — role={emp.role}, hours={display_hours}, "
                f"revenue={display_revenue}, direct_comm={direct_comm}, override_comm={override_comm}"
            )

            direct_commission += direct_comm
            override_commission += override_comm
            report_data.append({
                "project_number": project.project_id,
                "project_name": project.name,
                "hours_billed": display_hours,
                "bill_rate": staff.hourly_rate or 0,
                "direct_commission": direct_comm,
                "override_commission": override_comm,
                "total_commission": direct_comm + override_comm,
            })

    app.logger.info(
        f"Index report generated — direct_commission={direct_commission}, "
        f"override_commission={override_commission}, total={direct_commission + override_commission}"
    )

    export_url = None
    if employee and report_data:
        export_url = url_for(
            'commission_export',
            employee_id=employee.id,
            date_from=form.date_from.data.strftime('%Y-%m-%d') if form.date_from.data else '',
            date_to=form.date_to.data.strftime('%Y-%m-%d') if form.date_to.data else '',
        )

    return render_template(
        'index.html',
        form=form,
        report_data=report_data,
        employee=employee,
        direct_commission=direct_commission,
        override_commission=override_commission,
        total_commission=direct_commission + override_commission,
        export_url=export_url,
    )


@app.route('/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard():
    app.logger.debug(f"Dashboard accessed by {current_user.email}")

    form = DateRangeForm()

    date_from = form.date_from.data
    date_to = form.date_to.data
    app.logger.info(f"Dashboard date range filter: {date_from} to {date_to}")

    projects = Project.query.options(
        joinedload(Project.project_staff).joinedload(ProjectStaff.employee)
    ).filter_by(company_id=current_user.company_id).all()

    project_ids = [p.id for p in projects]

    hours_query = HoursEntry.query.filter(HoursEntry.project_id.in_(project_ids)) if project_ids else HoursEntry.query.filter(False)
    if date_from:
        hours_query = hours_query.filter(HoursEntry.date >= date_from)
    if date_to:
        hours_query = hours_query.filter(HoursEntry.date <= date_to)
    all_hours = hours_query.all()

    app.logger.debug(f"Fetched {len(all_hours)} hours entries for dashboard")

    # Map hours by project
    project_hours_map = {}
    for entry in all_hours:
        project_hours_map.setdefault(entry.project_id, []).append(entry)

    employee_summary = {}

    for project in projects:
        hours_entries = project_hours_map.get(project.id, [])
        staff_map = {ps.employee_id: ps for ps in project.project_staff}

        emp_hours = {}
        emp_revenue = {}
        total_project_revenue = 0

        for entry in hours_entries:
            staff = staff_map.get(entry.employee_id)
            if not staff:
                continue

            emp_hours[entry.employee_id] = emp_hours.get(entry.employee_id, 0) + entry.hours_worked
            revenue = entry.hours_worked * (staff.hourly_rate or 0)
            emp_revenue[entry.employee_id] = emp_revenue.get(entry.employee_id, 0) + revenue
            total_project_revenue += revenue

        # Associate-specific aggregates for non-associate commission basis
        total_assoc_hours = sum(
            emp_hours.get(emp_id, 0)
            for emp_id, s in staff_map.items()
            if s.employee and s.employee.role.lower() == 'associate'
        )
        total_assoc_revenue = sum(
            emp_revenue.get(emp_id, 0)
            for emp_id, s in staff_map.items()
            if s.employee and s.employee.role.lower() == 'associate'
        )
        total_project_hours = sum(emp_hours.get(emp_id, 0) for emp_id in staff_map)

        for emp_id, staff in staff_map.items():
            employee = staff.employee
            hours = emp_hours.get(emp_id, 0)
            revenue = emp_revenue.get(emp_id, 0)
            commission = 0.0

            role = employee.role.lower()
            if role == 'associate':
                direct_comm = revenue * (staff.commission_percentage or 0) / 100
                override_comm = 0
                display_hours = hours
                display_revenue = revenue
            elif role == 'director':
                direct_comm = revenue * (staff.commission_percentage or 0) / 100
                override_comm = (total_project_revenue * employee.override_percentage / 100) if employee.override_percentage else 0
                display_hours = total_project_hours
                display_revenue = total_project_revenue
            else:
                direct_comm = total_assoc_revenue * (staff.commission_percentage or 0) / 100
                override_comm = 0
                display_hours = total_assoc_hours
                display_revenue = total_assoc_revenue

            commission = direct_comm + override_comm

            if emp_id not in employee_summary:
                employee_summary[emp_id] = {
                    'employee': employee,
                    'total_hours': 0.0,
                    'total_revenue': 0.0,
                    'total_commission': 0.0,
                    'notes': set(),
                    "by_project": [],
                    "by_type": {},
                }

            summary = employee_summary[emp_id]
            summary['total_hours'] += display_hours
            summary['total_revenue'] += display_revenue
            summary['total_commission'] += commission
            summary['notes'].add(employee.role)
            summary['by_project'].append({"project": project.name, "hours": display_hours, "revenue": display_revenue, "commission": commission})

        app.logger.debug(f"Processed project: {project.name}")

    report_data = []
    for summary in employee_summary.values():
        report_data.append({
            'employee': summary['employee'],
            'total_hours': round(summary['total_hours'], 2),
            'total_revenue': round(summary['total_revenue'], 2),
            'total_commission': round(summary['total_commission'], 2),
            'notes': ", ".join(summary['notes']),
            "by_project": summary['by_project'],
        })

    app.logger.info(f"Dashboard report generated with {len(report_data)} employees summarised")

    dashboard_export_url = url_for(
        'dashboard_export',
        date_from=date_from.strftime('%Y-%m-%d') if date_from else '',
        date_to=date_to.strftime('%Y-%m-%d') if date_to else '',
    )

    return render_template('dashboard.html', form=form, report_data=report_data, export_url=dashboard_export_url)


# Projects routes
@app.route('/projects')
@login_required
def projects_list():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str).strip()
    sort = request.args.get('sort', 'name', type=str)
    order = request.args.get('order', 'asc', type=str)

    app.logger.info(f"[Projects List] Fetching projects for company_id={current_user.company_id}, page={page}, search={search}, sort={sort}, order={order}")

    query = Project.query.filter_by(company_id=current_user.company_id)

    if search:
        query = query.filter(or_(
            Project.name.ilike(f'%{search}%'),
            Project.client.ilike(f'%{search}%'),
            Project.project_id.ilike(f'%{search}%'),
        ))

    sort_columns = {
        'name': Project.name,
        'client': Project.client,
        'project_id': Project.project_id,
        'allocated_hours': Project.total_allocated_hours,
        'extra_hours': Project.extra_hours,
    }
    sort_col = sort_columns.get(sort, Project.name)
    query = query.order_by(sort_col.desc() if order == 'desc' else sort_col.asc())

    projects = get_paginated_query(query, page, per_page)

    # Batch pre-compute hours_worked and revenue to avoid N+1
    if projects.items:
        project_ids = [p.id for p in projects.items]

        hours_stats = db.session.query(
            HoursEntry.project_id,
            func.coalesce(func.sum(HoursEntry.hours_worked), 0),
        ).filter(HoursEntry.project_id.in_(project_ids)).group_by(HoursEntry.project_id).all()
        hours_map = {row[0]: row[1] for row in hours_stats}

        revenue_stats = db.session.query(
            HoursEntry.project_id,
            func.coalesce(func.sum(HoursEntry.hours_billed * Employee.hourly_rate), 0),
        ).join(Employee).filter(HoursEntry.project_id.in_(project_ids)).group_by(HoursEntry.project_id).all()
        revenue_map = {row[0]: row[1] for row in revenue_stats}

        for p in projects.items:
            p._prefetched_hours_worked = hours_map.get(p.id, 0)
            p._prefetched_revenue = revenue_map.get(p.id, 0)

    app.logger.debug(f"[Projects List] Retrieved {len(projects.items)} projects on page {page}")
    return render_template('projects/list.html', projects=projects, search=search, sort=sort, order=order)


@app.route('/projects/new', methods=['GET', 'POST'])
@login_required
def projects_new():
    app.logger.info(f"[Projects New] Request method={request.method}")
    form = ProjectForm()
    if form.validate_on_submit():
        app.logger.debug(f"[Projects New] Creating project '{form.name.data}' for company_id={current_user.company_id}")
        project = Project(
            name=form.name.data,
            project_id=form.project_id.data,
            client=form.client.data,
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            total_allocated_hours=form.total_allocated_hours.data,
            extra_hours=form.extra_hours.data,
            company_id=current_user.company_id
        )
        db.session.add(project)
        db.session.commit()
        app.logger.info(f"[Projects New] Project '{project.name}' created successfully with id={project.id}")
        flash('Project created successfully!', 'success')
        return redirect(url_for('projects_list'))
    return render_template('projects/form.html', form=form, title='New Project')


@app.route('/projects/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def projects_edit(id):
    app.logger.info(f"[Projects Edit] Editing project id={id} for company_id={current_user.company_id}")
    project = Project.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    form = ProjectForm(obj=project)
    if form.validate_on_submit():
        app.logger.debug(f"[Projects Edit] Updating project id={project.id}")
        form.populate_obj(project)
        db.session.commit()
        app.logger.info(f"[Projects Edit] Project id={project.id} updated successfully")
        flash('Project updated successfully!', 'success')
        return redirect(url_for('projects_list'))
    return render_template('projects/form.html', form=form, title='Edit Project', project=project)


@app.route('/projects/<int:id>/delete', methods=['POST'])
@login_required
def projects_delete(id):
    app.logger.info(f"[Projects Delete] Deleting project id={id} for company_id={current_user.company_id}")
    project = Project.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    db.session.delete(project)
    db.session.commit()
    app.logger.info(f"[Projects Delete] Project id={id} deleted successfully")
    flash('Project deleted successfully!', 'success')
    return redirect(url_for('projects_list'))


@app.route('/projects/<int:id>')
@login_required
def projects_detail(id):
    app.logger.info(f"[Projects Detail] Fetching details for project id={id} (company_id={current_user.company_id})")
    project = Project.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()

    hour_entries = HoursEntry.query.options(
        joinedload(HoursEntry.employee),
    ).filter_by(project_id=project.id).order_by(
        HoursEntry.date.desc(), HoursEntry.created_at.desc()
    ).all()

    # Pre-fetch ProjectStaff and pre-compute commission_earned to avoid N+1
    if hour_entries:
        pairs = set((e.employee_id, e.project_id) for e in hour_entries)
        pair_filters = [
            and_(ProjectStaff.employee_id == eid, ProjectStaff.project_id == pid)
            for eid, pid in pairs
        ]
        all_staff = ProjectStaff.query.filter(or_(*pair_filters)).all() if pair_filters else []
        staff_map = {(s.employee_id, s.project_id): s for s in all_staff}
        for entry in hour_entries:
            entry._prefetched_project_staff = staff_map.get((entry.employee_id, entry.project_id))

        # Compute project-level aggregates in Python from already-loaded entries
        project_revenue = sum(
            e.hours_billed * (e.employee.hourly_rate or 0) for e in hour_entries
        )
        associate_emp_ids = {
            e.employee_id for e in hour_entries if e.employee.role.lower() == 'associate'
        }
        associate_revenue = sum(
            e.hours_billed * (e.employee.hourly_rate or 0)
            for e in hour_entries if e.employee_id in associate_emp_ids
        )
        for entry in hour_entries:
            staff = entry._prefetched_project_staff
            if not staff or not entry.employee:
                entry._prefetched_commission_earned = 0.0
                continue
            commission_pct = (staff.commission_percentage or 0) / 100
            hourly_rate = entry.employee.hourly_rate or 0
            role = entry.employee.role.lower()
            if role == 'associate':
                comm = entry.hours_worked * hourly_rate * commission_pct
            elif role == 'director':
                comm = entry.hours_worked * hourly_rate * commission_pct
                comm += (project_revenue * entry.employee.override_percentage / 100) if entry.employee.override_percentage else 0
            else:
                comm = associate_revenue * commission_pct
            entry._prefetched_commission_earned = comm

    app.logger.debug(f"[Projects Detail] Retrieved {len(hour_entries)} hour entries for project id={project.id}")

    # Per-employee summary for Hours Summary table
    emp_summary = {}
    for entry in hour_entries:
        eid = entry.employee_id
        if eid not in emp_summary:
            emp_summary[eid] = {
                'name': entry.employee.name,
                'role': entry.employee.role,
                'hours_billed': 0.0,
                'revenue': 0.0,
                'commission': 0.0,
            }
        emp_summary[eid]['hours_billed'] += entry.hours_billed or 0
        emp_summary[eid]['revenue'] += entry.revenue_generated or 0
        emp_summary[eid]['commission'] += entry.commission_earned or 0
    hours_summary = sorted(emp_summary.values(), key=lambda x: x['name'])

    return render_template('projects/detail.html', project=project, hour_entries=hour_entries, hours_summary=hours_summary)

# Employees routes
@app.route('/employees')
@login_required
def employees_list():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str).strip()
    sort = request.args.get('sort', 'name', type=str)
    order = request.args.get('order', 'asc', type=str)

    app.logger.info(f"[Employees List] Fetching employees for company_id={current_user.company_id}, page={page}, search={search}, sort={sort}, order={order}")

    query = Employee.query.options(
        subqueryload(Employee.project_staff),
    ).filter_by(company_id=current_user.company_id)

    if search:
        query = query.filter(or_(
            Employee.name.ilike(f'%{search}%'),
            Employee.role.ilike(f'%{search}%'),
        ))

    sort_columns = {
        'name': Employee.name,
        'role': Employee.role,
        'hourly_rate': Employee.hourly_rate,
        'override_percentage': Employee.override_percentage,
    }
    sort_col = sort_columns.get(sort, Employee.name)
    query = query.order_by(sort_col.desc() if order == 'desc' else sort_col.asc())

    employees = get_paginated_query(query, page, per_page)

    # Batch pre-compute total_hours_worked to avoid N+1
    if employees.items:
        emp_ids = [e.id for e in employees.items]
        hours_stats = db.session.query(
            HoursEntry.employee_id,
            func.coalesce(func.sum(HoursEntry.hours_worked), 0),
        ).filter(HoursEntry.employee_id.in_(emp_ids)).group_by(HoursEntry.employee_id).all()
        hours_map = {row[0]: row[1] for row in hours_stats}

        for e in employees.items:
            e._prefetched_hours_worked = hours_map.get(e.id, 0)

    app.logger.debug(f"[Employees List] Retrieved {len(employees.items)} employees on page {page}")
    return render_template('employees/list.html', employees=employees, search=search, sort=sort, order=order)


@app.route('/employees/new', methods=['GET', 'POST'])
@login_required
def employees_new():
    app.logger.info(f"[Employees New] Request method={request.method}")
    form = EmployeeForm()
    if form.validate_on_submit():
        app.logger.debug(f"[Employees New] Creating employee '{form.name.data}' for company_id={current_user.company_id}")
        print(form.role.data,form.override_percentage.data, form.override_percentage.data if form.role.data == 'Director' else 0.0, "override_percentage")
        employee = Employee(
            name=form.name.data,
            role=form.role.data,
            hourly_rate=form.hourly_rate.data,
            override_percentage=(form.override_percentage.data if form.override_percentage.data is not None else 2.0) if form.role.data == 'Director' else 0.0,
            company_id=current_user.company_id
        )
        db.session.add(employee)
        db.session.commit()
        app.logger.info(f"[Employees New] Employee '{employee.name}' created successfully with id={employee.id}")
        flash('Employee created successfully!', 'success')
        return redirect(url_for('employees_list'))
    return render_template('employees/form.html', form=form, title='New Employee')


@app.route('/employees/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def employees_edit(id):
    app.logger.info(f"[Employees Edit] Editing employee id={id} for company_id={current_user.company_id}")
    employee = Employee.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    form = EmployeeForm(obj=employee)
    if form.validate_on_submit():
        app.logger.debug(f"[Employees Edit] Updating employee id={employee.id}")
        form.populate_obj(employee)
        if employee.role == "Director":
            employee.override_percentage = form.override_percentage.data if form.override_percentage.data is not None else 2.0
        else:
            employee.override_percentage = 0.0
        db.session.commit()
        app.logger.info(f"[Employees Edit] Employee id={employee.id} updated successfully")
        flash('Employee updated successfully!', 'success')
        return redirect(url_for('employees_list'))
    return render_template('employees/form.html', form=form, title='Edit Employee', employee=employee)


@app.route('/employees/<int:id>/delete', methods=['POST'])
@login_required
def employees_delete(id):
    app.logger.info(f"[Employees Delete] Deleting employee id={id} for company_id={current_user.company_id}")
    employee = Employee.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    db.session.delete(employee)
    db.session.commit()
    app.logger.info(f"[Employees Delete] Employee id={id} deleted successfully")
    flash('Employee deleted successfully!', 'success')
    return redirect(url_for('employees_list'))

@app.route("/employees/<int:id>/rate")
@login_required
def get_employee_rate(id):
    employee = Employee.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    return {"hourly_rate": employee.hourly_rate}

# Project Staff routes
@app.route('/project-staff')
@login_required
def project_staff_list():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str).strip()
    sort = request.args.get('sort', 'created_at', type=str)
    order = request.args.get('order', 'desc', type=str)

    app.logger.info(f"[ProjectStaff List] Fetching for company_id={current_user.company_id}, page={page}, search={search}, sort={sort}, order={order}")

    query = ProjectStaff.query.join(Project).join(Employee).options(
        contains_eager(ProjectStaff.project),
        contains_eager(ProjectStaff.employee),
    ).filter(Project.company_id == current_user.company_id)

    if search:
        query = query.filter(or_(
            Employee.name.ilike(f'%{search}%'),
            Project.name.ilike(f'%{search}%'),
        ))

    sort_columns = {
        'employee': Employee.name,
        'project': Project.name,
        'hourly_rate': ProjectStaff.hourly_rate,
        'commission_percentage': ProjectStaff.commission_percentage,
        'created_at': ProjectStaff.created_at,
    }
    sort_col = sort_columns.get(sort, ProjectStaff.created_at)
    query = query.order_by(sort_col.desc() if order == 'desc' else sort_col.asc())

    project_staff = get_paginated_query(query, page, per_page)
    app.logger.debug(f"[ProjectStaff List] Retrieved {len(project_staff.items)} assignments on page {page}")
    return render_template('project_staff/list.html', project_staff=project_staff, search=search, sort=sort, order=order)


@app.route('/project-staff/new', methods=['GET', 'POST'])
@login_required
def project_staff_new():
    app.logger.info(f"[ProjectStaff New] Request method={request.method}")
    form = ProjectStaffForm()
    if form.validate_on_submit():
        try:
            app.logger.debug(f"[ProjectStaff New] Assigning employee_id={form.employee_id.data} to project_id={form.project_id.data}")
            project_staff = ProjectStaff(
                company_id=current_user.company_id,
                employee_id=form.employee_id.data,
                project_id=form.project_id.data,
                hourly_rate=form.hourly_rate.data,
                commission_percentage=form.commission_percentage.data,
            )
            db.session.add(project_staff)
            db.session.commit()
            app.logger.info(f"[ProjectStaff New] Assignment created with id={project_staff.id}")
            flash('Employee assigned to project successfully!', 'success')
            return redirect(url_for('project_staff_list'))
        except IntegrityError as e:
            db.session.rollback()
            app.logger.warning(f"[ProjectStaff New] IntegrityError: {str(e)}")
            flash('This employee is already assigned to the selected project.', 'error')
    return render_template('project_staff/form.html', form=form, title='Assign Employee to Project')


@app.route('/project-staff/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def project_staff_edit(id):
    app.logger.info(f"[ProjectStaff Edit] Editing assignment id={id} for company_id={current_user.company_id}")
    project_staff = ProjectStaff.query.join(Project).filter(
        ProjectStaff.id == id, 
        Project.company_id == current_user.company_id
    ).first_or_404()
    form = ProjectStaffForm(obj=project_staff)
    if form.validate_on_submit():
        try:
            app.logger.debug(f"[ProjectStaff Edit] Updating assignment id={project_staff.id}")
            form.populate_obj(project_staff)
            db.session.commit()
            app.logger.info(f"[ProjectStaff Edit] Assignment id={project_staff.id} updated successfully")
            flash('Project assignment updated successfully!', 'success')
            return redirect(url_for('project_staff_list'))
        except IntegrityError as e:
            db.session.rollback()
            app.logger.warning(f"[ProjectStaff New] IntegrityError: {str(e)}")
            flash('This employee is already assigned to the selected project.', 'error')
    return render_template('project_staff/form.html', form=form, title='Edit Project Assignment', project_staff=project_staff)


@app.route('/project-staff/<int:id>/delete', methods=['POST'])
@login_required
def project_staff_delete(id):
    app.logger.info(f"[ProjectStaff Delete] Deleting assignment id={id} for company_id={current_user.company_id}")
    project_staff = ProjectStaff.query.join(Project).filter(
        ProjectStaff.id == id, 
        Project.company_id == current_user.company_id
    ).first_or_404()
    db.session.delete(project_staff)
    db.session.commit()
    app.logger.info(f"[ProjectStaff Delete] Assignment id={id} deleted successfully")
    flash('Project assignment removed successfully!', 'success')
    return redirect(url_for('project_staff_list'))


# Hours Entry routes
@app.route('/hours')
@login_required
def hours_list():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    search = request.args.get('search', '', type=str).strip()
    sort = request.args.get('sort', 'date', type=str)
    order = request.args.get('order', 'desc', type=str)

    app.logger.info(f"[Hours List] Fetching hours for company_id={current_user.company_id}, page={page}, search={search}, sort={sort}, order={order}")

    query = HoursEntry.query.join(Project).join(Employee).options(
        contains_eager(HoursEntry.project),
        contains_eager(HoursEntry.employee),
    ).filter(Project.company_id == current_user.company_id)

    if search:
        query = query.filter(or_(
            Employee.name.ilike(f'%{search}%'),
            Project.name.ilike(f'%{search}%'),
            HoursEntry.description.ilike(f'%{search}%'),
        ))

    sort_columns = {
        'date': HoursEntry.date,
        'employee': Employee.name,
        'project': Project.name,
        'hours_billed': HoursEntry.hours_billed,
    }
    sort_col = sort_columns.get(sort, HoursEntry.date)
    query = query.order_by(sort_col.desc() if order == 'desc' else sort_col.asc())

    hours = get_paginated_query(query, page, per_page)

    # Pre-fetch ProjectStaff and pre-compute commission_earned to avoid N+1
    if hours.items:
        pairs = set((e.employee_id, e.project_id) for e in hours.items)
        pair_filters = [
            and_(ProjectStaff.employee_id == eid, ProjectStaff.project_id == pid)
            for eid, pid in pairs
        ]
        all_staff = ProjectStaff.query.filter(or_(*pair_filters)).all() if pair_filters else []
        staff_map = {(s.employee_id, s.project_id): s for s in all_staff}
        for entry in hours.items:
            entry._prefetched_project_staff = staff_map.get((entry.employee_id, entry.project_id))

        # Batch aggregate project and associate revenue for all projects on this page
        page_project_ids = list({e.project_id for e in hours.items})
        proj_rev_rows = db.session.query(
            HoursEntry.project_id,
            func.coalesce(func.sum(HoursEntry.hours_billed * Employee.hourly_rate), 0)
        ).join(Employee, HoursEntry.employee_id == Employee.id).filter(
            HoursEntry.project_id.in_(page_project_ids)
        ).group_by(HoursEntry.project_id).all()
        proj_rev_map = {row[0]: row[1] for row in proj_rev_rows}

        assoc_rev_rows = db.session.query(
            HoursEntry.project_id,
            func.coalesce(func.sum(HoursEntry.hours_billed * Employee.hourly_rate), 0)
        ).join(Employee, HoursEntry.employee_id == Employee.id).filter(
            HoursEntry.project_id.in_(page_project_ids),
            Employee.role.ilike('associate')
        ).group_by(HoursEntry.project_id).all()
        assoc_rev_map = {row[0]: row[1] for row in assoc_rev_rows}

        for entry in hours.items:
            staff = staff_map.get((entry.employee_id, entry.project_id))
            if not staff or not entry.employee:
                entry._prefetched_commission_earned = 0.0
                continue
            commission_pct = (staff.commission_percentage or 0) / 100
            hourly_rate = entry.employee.hourly_rate or 0
            role = entry.employee.role.lower()
            project_revenue = proj_rev_map.get(entry.project_id, 0)
            associate_revenue = assoc_rev_map.get(entry.project_id, 0)
            if role == 'associate':
                comm = entry.hours_worked * hourly_rate * commission_pct
            elif role == 'director':
                comm = entry.hours_worked * hourly_rate * commission_pct
                comm += (project_revenue * entry.employee.override_percentage / 100) if entry.employee.override_percentage else 0
            else:
                comm = associate_revenue * commission_pct
            entry._prefetched_commission_earned = comm

    app.logger.debug(f"[Hours List] Retrieved {len(hours.items)} entries on page {page}")
    return render_template('hours/list.html', hours=hours, search=search, sort=sort, order=order)


@app.route('/hours/new', methods=['GET', 'POST'])
@login_required
def hours_new():
    app.logger.info(f"[Hours New] Request method={request.method}")
    form = HoursEntryForm()
    if form.validate_on_submit():
        app.logger.debug(f"[Hours New] Adding hours for employee_id={form.employee_id.data} on project_id={form.project_id.data}")
        hours_entry = HoursEntry(
            employee_id=form.employee_id.data,
            project_id=form.project_id.data,
            date=form.date.data,
            hours_worked=form.hours_billed.data,  # same for now
            hours_billed=form.hours_billed.data,
            description=form.description.data,
            company_id=current_user.company_id
        )
        db.session.add(hours_entry)
        db.session.commit()
        app.logger.info(f"[Hours New] Hours entry created with id={hours_entry.id}")
        flash('Hours entry created successfully!', 'success')
        return redirect(url_for('hours_list'))
    return render_template('hours/form.html', form=form, title='New Hours Entry')


@app.route('/hours/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def hours_edit(id):
    app.logger.info(f"[Hours Edit] Editing hours entry id={id} for company_id={current_user.company_id}")
    hours_entry = HoursEntry.query.join(Project).filter(
        HoursEntry.id == id, 
        Project.company_id == current_user.company_id
    ).first_or_404()
    form = HoursEntryForm(obj=hours_entry)
    if form.validate_on_submit():
        app.logger.debug(f"[Hours Edit] Updating hours entry id={hours_entry.id}")
        form.populate_obj(hours_entry)
        db.session.commit()
        app.logger.info(f"[Hours Edit] Hours entry id={hours_entry.id} updated successfully")
        flash('Hours entry updated successfully!', 'success')
        return redirect(url_for('hours_list'))
    return render_template('hours/form.html', form=form, title='Edit Hours Entry', hours_entry=hours_entry)


@app.route('/hours/<int:id>/delete', methods=['POST'])
@login_required
def hours_delete(id):
    redirect_back = request.args.get("redirect", "0") == "1"
    print(redirect_back)
    app.logger.info(f"[Hours Delete] Deleting hours entry id={id} for company_id={current_user.company_id}")
    hours_entry = HoursEntry.query.join(Project).filter(
        HoursEntry.id == id, 
        Project.company_id == current_user.company_id
    ).first_or_404()
    db.session.delete(hours_entry)
    db.session.commit()
    app.logger.info(f"[Hours Delete] Hours entry id={id} deleted successfully")
    flash('Hours entry deleted successfully!', 'success')
    if redirect_back:
        return redirect(url_for('hours_list'))
    else:
        return redirect(url_for('projects_detail', id=hours_entry.project_id))
    
# Error handlers
@app.errorhandler(404)
def not_found(error):
    app.logger.warning(f"[404 Not Found] URL: {request.path} | Method: {request.method} | Error: {error}")
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    app.logger.error(f"[500 Internal Error] URL: {request.path} | Method: {request.method} | Error: {error}", exc_info=True)
    db.session.rollback()
    return render_template('500.html'), 500


# ─── Excel Export / Import / Sample Routes ────────────────────────────────────

# --- Projects ---

@app.route('/projects/export')
@login_required
def projects_export():
    projects = Project.query.filter_by(company_id=current_user.company_id).order_by(Project.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Projects'
    ws.append(['Project ID', 'Name', 'Client', 'Start Date', 'End Date', 'Total Allocated Hours', 'Extra Hours'])
    _style_header_row(ws)
    for p in projects:
        ws.append([
            p.project_id, p.name, p.client,
            p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
            p.end_date.strftime('%Y-%m-%d') if p.end_date else '',
            float(p.total_allocated_hours or 0),
            float(p.extra_hours or 0),
        ])
    _auto_width(ws)
    return _make_excel_response(wb, 'projects.xlsx')


@app.route('/projects/sample')
@login_required
def projects_sample():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Projects'
    ws.append(['Project ID', 'Name', 'Client', 'Start Date', 'End Date', 'Total Allocated Hours', 'Extra Hours'])
    _style_header_row(ws)
    ws.append(['P001', 'Sample Project', 'Acme Corp', '2024-01-01', '2024-12-31', 1000.0, 100.0])
    _auto_width(ws)
    return _make_excel_response(wb, 'projects_sample.xlsx')


@app.route('/projects/import', methods=['POST'])
@login_required
def projects_import():
    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('projects_list'))
    try:
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created, errors = 0, []
        for i, row in enumerate(rows, start=2):
            row = list(row) + [None] * 7
            project_id, name, client, start_date, end_date, allocated_hours, extra_hours = row[:7]
            if not project_id or not name or not client:
                errors.append(f'Row {i}: Missing required fields (Project ID, Name, Client).')
                continue
            try:
                sd = datetime.strptime(str(start_date).strip(), '%Y-%m-%d').date() if start_date else None
                ed = datetime.strptime(str(end_date).strip(), '%Y-%m-%d').date() if end_date else None
                db.session.add(Project(
                    project_id=str(project_id).strip(), name=str(name).strip(), client=str(client).strip(),
                    start_date=sd, end_date=ed,
                    total_allocated_hours=float(allocated_hours or 0),
                    extra_hours=float(extra_hours or 0),
                    company_id=current_user.company_id,
                ))
                created += 1
            except Exception as e:
                errors.append(f'Row {i}: {e}')
        db.session.commit()
        if created:
            flash(f'Imported {created} project(s) successfully.', 'success')
        for err in errors[:5]:
            flash(err, 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Import failed: {e}', 'error')
    return redirect(url_for('projects_list'))


# --- Employees ---

@app.route('/employees/export')
@login_required
def employees_export():
    employees = Employee.query.filter_by(company_id=current_user.company_id).order_by(Employee.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employees'
    ws.append(['Name', 'Role', 'Hourly Rate', 'Override Percentage'])
    _style_header_row(ws)
    for e in employees:
        ws.append([e.name, e.role, float(e.hourly_rate or 0), float(e.override_percentage or 0)])
    _auto_width(ws)
    return _make_excel_response(wb, 'employees.xlsx')


@app.route('/employees/sample')
@login_required
def employees_sample():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Employees'
    ws.append(['Name', 'Role', 'Hourly Rate', 'Override Percentage'])
    _style_header_row(ws)
    ws.append(['Jane Smith', 'Associate', 75.0, 0.0])
    ws.append(['John Director', 'Director', 120.0, 2.0])
    ws.append(['Alice PM', 'PM', 90.0, 0.0])
    _auto_width(ws)
    return _make_excel_response(wb, 'employees_sample.xlsx')


@app.route('/employees/import', methods=['POST'])
@login_required
def employees_import():
    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('employees_list'))
    valid_roles = {'Associate', 'Director', 'PM', 'Analyst', 'Consultant'}
    try:
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created, errors = 0, []
        for i, row in enumerate(rows, start=2):
            row = list(row) + [None] * 4
            name, role, hourly_rate, override_pct = row[:4]
            if not name or not role:
                errors.append(f'Row {i}: Missing required fields (Name, Role).')
                continue
            role = str(role).strip()
            if role not in valid_roles:
                errors.append(f'Row {i}: Invalid role "{role}". Must be one of: {", ".join(valid_roles)}.')
                continue
            try:
                override = float(override_pct or 0) if role == 'Director' else 0.0
                db.session.add(Employee(
                    name=str(name).strip(), role=role,
                    hourly_rate=float(hourly_rate or 0), override_percentage=override,
                    company_id=current_user.company_id,
                ))
                created += 1
            except Exception as e:
                errors.append(f'Row {i}: {e}')
        db.session.commit()
        if created:
            flash(f'Imported {created} employee(s) successfully.', 'success')
        for err in errors[:5]:
            flash(err, 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Import failed: {e}', 'error')
    return redirect(url_for('employees_list'))


# --- Project Staff ---

@app.route('/project-staff/export')
@login_required
def project_staff_export():
    assignments = ProjectStaff.query.join(Project).join(Employee).options(
        contains_eager(ProjectStaff.project), contains_eager(ProjectStaff.employee),
    ).filter(Project.company_id == current_user.company_id).order_by(Employee.name).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Project Staff'
    ws.append(['Employee Name', 'Project Name', 'Commission Percentage', 'Hourly Rate'])
    _style_header_row(ws)
    for a in assignments:
        ws.append([a.employee.name, a.project.name, float(a.commission_percentage or 0), float(a.hourly_rate or 0)])
    _auto_width(ws)
    return _make_excel_response(wb, 'project_staff.xlsx')


@app.route('/project-staff/sample')
@login_required
def project_staff_sample():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Project Staff'
    ws.append(['Employee Name', 'Project Name', 'Commission Percentage', 'Hourly Rate'])
    _style_header_row(ws)
    ws.append(['Jane Smith', 'Sample Project', 5.0, 75.0])
    _auto_width(ws)
    return _make_excel_response(wb, 'project_staff_sample.xlsx')


@app.route('/project-staff/import', methods=['POST'])
@login_required
def project_staff_import():
    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('project_staff_list'))
    try:
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created, errors = 0, []
        for i, row in enumerate(rows, start=2):
            row = list(row) + [None] * 4
            emp_name, proj_name, commission_pct, hourly_rate = row[:4]
            if not emp_name or not proj_name:
                errors.append(f'Row {i}: Missing required fields (Employee Name, Project Name).')
                continue
            emp = Employee.query.filter_by(name=str(emp_name).strip(), company_id=current_user.company_id).first()
            if not emp:
                errors.append(f'Row {i}: Employee "{emp_name}" not found.')
                continue
            proj = Project.query.filter_by(name=str(proj_name).strip(), company_id=current_user.company_id).first()
            if not proj:
                errors.append(f'Row {i}: Project "{proj_name}" not found.')
                continue
            try:
                db.session.add(ProjectStaff(
                    employee_id=emp.id, project_id=proj.id,
                    commission_percentage=float(commission_pct or 0),
                    hourly_rate=float(hourly_rate or emp.hourly_rate or 0),
                    company_id=current_user.company_id,
                ))
                created += 1
            except Exception as e:
                errors.append(f'Row {i}: {e}')
        db.session.commit()
        if created:
            flash(f'Imported {created} assignment(s) successfully.', 'success')
        for err in errors[:5]:
            flash(err, 'error')
    except IntegrityError:
        db.session.rollback()
        flash('Some assignments already exist and were skipped.', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Import failed: {e}', 'error')
    return redirect(url_for('project_staff_list'))


# --- Hours ---

@app.route('/hours/export')
@login_required
def hours_export():
    entries = HoursEntry.query.join(Project).join(Employee).options(
        contains_eager(HoursEntry.project), contains_eager(HoursEntry.employee),
    ).filter(Project.company_id == current_user.company_id).order_by(HoursEntry.date.desc()).all()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hours'
    ws.append(['Employee Name', 'Project Name', 'Date', 'Hours Billed', 'Description'])
    _style_header_row(ws)
    for e in entries:
        ws.append([
            e.employee.name, e.project.name,
            e.date.strftime('%Y-%m-%d') if e.date else '',
            float(e.hours_billed or 0),
            e.description or '',
        ])
    _auto_width(ws)
    return _make_excel_response(wb, 'hours.xlsx')


@app.route('/hours/sample')
@login_required
def hours_sample():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hours'
    ws.append(['Employee Name', 'Project Name', 'Date', 'Hours Billed', 'Description'])
    _style_header_row(ws)
    ws.append(['Jane Smith', 'Sample Project', '2024-01-15', 8.0, 'Design review and implementation'])
    _auto_width(ws)
    return _make_excel_response(wb, 'hours_sample.xlsx')


@app.route('/hours/import', methods=['POST'])
@login_required
def hours_import():
    f = request.files.get('file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('hours_list'))
    try:
        wb = load_workbook(f, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created, errors = 0, []
        for i, row in enumerate(rows, start=2):
            row = list(row) + [None] * 5
            emp_name, proj_name, date_val, hours_billed, description = row[:5]
            if not emp_name or not proj_name or not date_val:
                errors.append(f'Row {i}: Missing required fields (Employee Name, Project Name, Date).')
                continue
            emp = Employee.query.filter_by(name=str(emp_name).strip(), company_id=current_user.company_id).first()
            if not emp:
                errors.append(f'Row {i}: Employee "{emp_name}" not found.')
                continue
            proj = Project.query.filter_by(name=str(proj_name).strip(), company_id=current_user.company_id).first()
            if not proj:
                errors.append(f'Row {i}: Project "{proj_name}" not found.')
                continue
            if not ProjectStaff.query.filter_by(employee_id=emp.id, project_id=proj.id).first():
                errors.append(f'Row {i}: Employee "{emp_name}" is not assigned to project "{proj_name}".')
                continue
            try:
                if isinstance(date_val, str):
                    entry_date = datetime.strptime(date_val.strip(), '%Y-%m-%d').date()
                else:
                    entry_date = date_val
                billed = float(hours_billed or 0)
                db.session.add(HoursEntry(
                    employee_id=emp.id, project_id=proj.id,
                    date=entry_date, hours_worked=billed, hours_billed=billed,
                    description=str(description or '').strip(),
                    company_id=current_user.company_id,
                ))
                created += 1
            except Exception as e:
                errors.append(f'Row {i}: {e}')
        db.session.commit()
        if created:
            flash(f'Imported {created} hours entry(s) successfully.', 'success')
        for err in errors[:5]:
            flash(err, 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Import failed: {e}', 'error')
    return redirect(url_for('hours_list'))


# --- Commission Report Export ---

@app.route('/commission/export')
@login_required
def commission_export():
    employee_id = request.args.get('employee_id', type=int)
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')

    if not employee_id or not date_from_str or not date_to_str:
        flash('Missing export parameters.', 'error')
        return redirect(url_for('index'))

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format.', 'error')
        return redirect(url_for('index'))

    employee = Employee.query.filter_by(id=employee_id, company_id=current_user.company_id).first_or_404()

    projects = Project.query.options(
        joinedload(Project.project_staff).joinedload(ProjectStaff.employee)
    ).join(ProjectStaff).filter(
        Project.company_id == current_user.company_id,
        ProjectStaff.employee_id == employee_id
    ).all()

    project_ids = [p.id for p in projects]
    all_project_hours = HoursEntry.query.filter(
        HoursEntry.project_id.in_(project_ids),
        HoursEntry.date >= date_from, HoursEntry.date <= date_to
    ).all() if project_ids else []

    project_emp_hours = {}
    for entry in all_project_hours:
        project_emp_hours.setdefault(entry.project_id, {})
        project_emp_hours[entry.project_id][entry.employee_id] = (
            project_emp_hours[entry.project_id].get(entry.employee_id, 0) + entry.hours_worked
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Commission Report'
    ws.append(['Project ID', 'Project Name', 'Hours Billed', 'Bill Rate', 'Direct Commission', 'Override Commission', 'Total Commission'])
    _style_header_row(ws)

    for project in projects:
        staff_map = {ps.employee_id: ps for ps in project.project_staff}
        staff = staff_map.get(employee_id)
        if not staff:
            continue
        emp_hours_map = project_emp_hours.get(project.id, {})
        hours = emp_hours_map.get(employee_id, 0)
        revenue = hours * (staff.hourly_rate or 0)
        total_project_revenue = sum(
            emp_hours_map.get(eid, 0) * (ps.hourly_rate or 0) for eid, ps in staff_map.items()
        )
        total_assoc_revenue = sum(
            emp_hours_map.get(eid, 0) * (ps.hourly_rate or 0)
            for eid, ps in staff_map.items()
            if ps.employee and ps.employee.role.lower() == 'associate'
        )
        role = employee.role.lower()
        if role == 'associate':
            direct_comm = revenue * (staff.commission_percentage or 0) / 100
            override_comm = 0
            display_hours = hours
        elif role == 'director':
            direct_comm = revenue * (staff.commission_percentage or 0) / 100
            override_comm = (total_project_revenue * employee.override_percentage / 100) if employee.override_percentage else 0
            display_hours = sum(emp_hours_map.get(eid, 0) for eid in staff_map)
        else:
            direct_comm = total_assoc_revenue * (staff.commission_percentage or 0) / 100
            override_comm = 0
            display_hours = sum(
                emp_hours_map.get(eid, 0) for eid, ps in staff_map.items()
                if ps.employee and ps.employee.role.lower() == 'associate'
            )
        ws.append([
            project.project_id, project.name,
            round(display_hours, 2), round(staff.hourly_rate or 0, 2),
            round(direct_comm, 2), round(override_comm, 2), round(direct_comm + override_comm, 2),
        ])

    _auto_width(ws)
    fname = f'commission_{employee.name.replace(" ", "_")}_{date_from_str}_{date_to_str}.xlsx'
    return _make_excel_response(wb, fname)


# --- Dashboard Export ---

@app.route('/dashboard/export')
@login_required
def dashboard_export():
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except ValueError:
        flash('Invalid date format.', 'error')
        return redirect(url_for('dashboard'))

    projects = Project.query.options(
        joinedload(Project.project_staff).joinedload(ProjectStaff.employee)
    ).filter_by(company_id=current_user.company_id).all()

    project_ids = [p.id for p in projects]
    hours_query = HoursEntry.query.filter(HoursEntry.project_id.in_(project_ids)) if project_ids else HoursEntry.query.filter(False)
    if date_from:
        hours_query = hours_query.filter(HoursEntry.date >= date_from)
    if date_to:
        hours_query = hours_query.filter(HoursEntry.date <= date_to)
    all_hours = hours_query.all()

    project_hours_map = {}
    for entry in all_hours:
        project_hours_map.setdefault(entry.project_id, []).append(entry)

    employee_summary = {}
    for project in projects:
        hours_entries = project_hours_map.get(project.id, [])
        staff_map = {ps.employee_id: ps for ps in project.project_staff}
        emp_hours, emp_revenue = {}, {}
        total_project_revenue = 0
        for entry in hours_entries:
            staff = staff_map.get(entry.employee_id)
            if not staff:
                continue
            emp_hours[entry.employee_id] = emp_hours.get(entry.employee_id, 0) + entry.hours_worked
            revenue = entry.hours_worked * (staff.hourly_rate or 0)
            emp_revenue[entry.employee_id] = emp_revenue.get(entry.employee_id, 0) + revenue
            total_project_revenue += revenue
        total_assoc_revenue = sum(
            emp_revenue.get(emp_id, 0) for emp_id, s in staff_map.items()
            if s.employee and s.employee.role.lower() == 'associate'
        )
        total_project_hours = sum(emp_hours.get(emp_id, 0) for emp_id in staff_map)
        for emp_id, staff in staff_map.items():
            emp = staff.employee
            hours = emp_hours.get(emp_id, 0)
            revenue = emp_revenue.get(emp_id, 0)
            role = emp.role.lower()
            if role == 'associate':
                commission = revenue * (staff.commission_percentage or 0) / 100
                display_hours, display_revenue = hours, revenue
            elif role == 'director':
                commission = (revenue * (staff.commission_percentage or 0) / 100 +
                              ((total_project_revenue * emp.override_percentage / 100) if emp.override_percentage else 0))
                display_hours, display_revenue = total_project_hours, total_project_revenue
            else:
                commission = total_assoc_revenue * (staff.commission_percentage or 0) / 100
                display_hours = sum(emp_hours.get(eid, 0) for eid, s in staff_map.items() if s.employee and s.employee.role.lower() == 'associate')
                display_revenue = total_assoc_revenue
            if emp_id not in employee_summary:
                employee_summary[emp_id] = {'employee': emp, 'total_hours': 0.0, 'total_revenue': 0.0, 'total_commission': 0.0}
            employee_summary[emp_id]['total_hours'] += display_hours
            employee_summary[emp_id]['total_revenue'] += display_revenue
            employee_summary[emp_id]['total_commission'] += commission

    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard'
    ws.append(['Employee', 'Role', 'Hours', 'Revenue', 'Commission'])
    _style_header_row(ws)
    for s in employee_summary.values():
        if s['total_revenue'] != 0 or s['total_commission'] != 0:
            ws.append([
                s['employee'].name, s['employee'].role,
                round(s['total_hours'], 2),
                round(s['total_revenue'], 2),
                round(s['total_commission'], 2),
            ])
    _auto_width(ws)
    return _make_excel_response(wb, 'dashboard.xlsx')


# Static routes
@app.route('/privacy-policy')
def privacy_policy():
    app.logger.info(f"[Privacy Policy] Page accessed from IP: {request.remote_addr}")
    return render_template('privacy_policy.html')

@app.route('/terms-conditions')
def terms_conditions():
    app.logger.info(f"[Terms & Conditions] Page accessed from IP: {request.remote_addr}")
    return render_template('terms_conditions.html')